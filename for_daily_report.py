import openpyxl
import re
import datetime
parsing_dict = {}  # словарь с распарсенными данными
parsing_dict_developments = {} # словарь с распарсенными данными событий обсуждения
summ_re = 0
status_for_del = ['обращение отклонено', 'поставлено на контроль', 'на рассмотрении модератора']
# раскладываем данные из файла в словарь по форме {id: данные для рассчета}
with open("for_daily_report.txt") as file:
    for line in file:
        key, *value = line.replace("\n", "").split("\t")
        parsing_dict[key] = value
#очищаем словарь от неиспользуемых статусов
parsing_dict = {k: v for k, v in parsing_dict.items() if v[1] not in status_for_del}
#перекладываем в словарь всего и повторных обращений
a = input('введите дату начала отчета ')
b = input('введите дату окончания отчета ')
parsing_dict1 = {k: v for k,v in parsing_dict.items() if datetime.datetime.strptime(a, "%d.%m.%Y") <= datetime.datetime.strptime(v[2], "%d.%m.%Y") <= datetime.datetime.strptime(b, "%d.%m.%Y")}#очищаем словарь под выбранные даты
re_appel = {v[9]: [str(parsing_dict1.values()).count(v[9])] for v in parsing_dict1.values()}
d = list(parsing_dict1.values())
for k in re_appel.keys():
    for v in range(len(d)):
        if k == d[v][9] and d[v][17] == 'пользователь не подтвердил решение вопроса':
            summ_re += 1
    re_appel[k] += [summ_re]
    summ_re = 0
#выводим разбивку по исполнителям и родительским обращениям
parents_isp = {str(v[19])[-8:-1]: [] for v in parsing_dict1.values() if v[17] == 'пользователь не подтвердил решение вопроса' and v[17] != ''}
parents_isp1 = {k: [v[15]] + [v[0]] + [v[3]] for k, v in parsing_dict.items() for d in parents_isp.keys() if k == d}
#делаем разбивку по источникам
source_of_appel = {v[16]: str(parsing_dict1.values()).count(v[16]) for v in parsing_dict1.values() if v[16] != ''}
#раскладываем данные из файла в словарь событий обсуждений
with open("sobytiya.txt") as file:
    for line in file:
        key, *value = line.replace("\n", "").split("\t")
        parsing_dict_developments[key] = value
parsing_dict_developments1 = {k: v for k,v in parsing_dict_developments.items() if a <= v[3] <= b}#очищаем словарь под выбранные даты
status_of_rejected = {v[10]: str(parsing_dict_developments1.values()).count(v[10]) for v in parsing_dict_developments1.values() if v[10] != ''} #делаем разбивку по статусам отклонения
source_of_appel.update(status_of_rejected)
#выясняем количество повторных по исполнителям
count = 0
isp_povt = {v[6]: [] for v in parsing_dict_developments1.values() if v[10] == 'Пользователь НЕ подтвердил решение вопроса в ходе звонка [5407016]'}
d = list(parsing_dict_developments1.values())
for k in isp_povt.keys():
    for v in range(len(d)):
        if k == d[v][6] and d[v][10] == 'Пользователь НЕ подтвердил решение вопроса в ходе звонка [5407016]':
            summ_re += 1
    isp_povt[k] = [summ_re]
    summ_re = 0
#детальная инфа по неподтвержденным звонкам
unconfirmed_calls = {re.sub(r'Обращение №', '', str(v[0])): [v[6]]  for k, v in parsing_dict_developments1.items() if v[10] == 'Пользователь НЕ подтвердил решение вопроса в ходе звонка [5407016]'}
unconfirmed_calls = {k: v + [parsing_dict[k][9]] + [parsing_dict[k][10]] + [parsing_dict[k][3]] for k, v in unconfirmed_calls.items()}
#выкладываем инфу в эксель
wb = openpyxl.Workbook()
wb.create_sheet(title = 'Отчет', index = 0)
wb.create_sheet(title = 'Детальная по повторным', index = 1)
sheet = wb['Отчет']
sheet.append(['Категория', 'Всего вопросов', 'Из них повторные'])
for k, v in re_appel.items():
    sheet.append([k, v[0], v[1]])
sheet2 = wb['Детальная по повторным']
sheet2.append(['Тема', 'Номер обращения', 'Исполнитель', 'Адрес'])
for k, v in parents_isp1.items():
    sheet2.append([v[1], k, v[0], v[2]])
wb.save('Принятые и повторные {} - {}.xlsx'.format(a, b)) 
wb = openpyxl.Workbook()
wb.create_sheet(title = 'Статистика по количеству', index = 0)
wb.create_sheet(title = 'Статистика по неподтвержденным', index = 1)
wb.create_sheet(title = 'Детальная информация по неподтвержденным', index = 2)
sheet = wb['Статистика по количеству']
stroka = list(source_of_appel.keys())
sheet.append([stroka[x] for x in range(len(stroka))])
stroka2 = list(source_of_appel.values())
sheet.append([stroka2[x] for x in range(len(stroka2))])
sheet2 = wb['Статистика по неподтвержденным']
sheet2.append(['Название МО', 'Количество неподтвержденных {} - {}'.format(a, b)])
for k, v in isp_povt.items():
    sheet2.append([k, v[0]])
sheet3 = wb['Детальная информация по неподтвержденным']
sheet3.append(['Тема', 'Номер обращения', 'Исполнитель', 'Адрес'])
for k, v in unconfirmed_calls.items():
    sheet3.append([v[1], k, v[0], v[2], v[3]])
wb.save('Итоги звонков {} - {}.xlsx'.format(a, b)) 
